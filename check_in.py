"""
程式名稱: checkin.py
版本: V1.2
更新內容:
1. 滿 9 小時工時檢核二次確認防呆機制 (未滿時暫緩寫入，提供 OK / 不OK 選擇)
2. 員工打卡紀錄隱私隔離 (僅呈現選定員工紀錄)
3. 自由 Key-in 修改/補登時間 (HH:MM 格式)
4. 繁體中文 / Bahasa Indonesia 雙語一鍵切換
"""

import streamlit as st
import pandas as pd
import sqlite3
import datetime
import io
import openpyxl
import re

# ==================== 系統常數與版本設定 ====================
APP_VERSION = "V1.2"
EMPLOYEES = ["OFW001(溫蒂)", "OFW002(都發)", "OFW003(菲娜)"]
DB_FILE = "attendance.db"

# ==================== 多語系字典 (雙語對照) ====================
TRANSLATIONS = {
    "zh": {
        "title": "外籍移工出勤打卡系統",
        "date_mgmt": "📅 日期與管理",
        "select_date": "選擇打卡日期",
        "current_view_date": "目前檢視日期",
        "export_title": "📤 匯出打卡紀錄",
        "export_range": "匯出範圍",
        "range_day": "當日紀錄",
        "range_month": "當月紀錄",
        "range_all": "全部紀錄",
        "download_excel": "📥 下載 Excel (匯入格式)",
        "no_export_data": "該範圍尚無打卡資料可供匯出。",
        "sys_version": "系統版本",
        "lang_label": "🌐 語言切換 (Bahasa)",
        "step1_select_emp": "1. 請選擇員工編號 (Pilih ID Karyawan)",
        "current_selected": "目前已選擇員工",
        "step2_clock": "2. 立即打卡 (Clock In / Out)",
        "clock_in": "🟢 上班 (Masuk)",
        "clock_out": "🔴 下班 (Pulang)",
        "clock_in_success": "上班打卡成功！時間",
        "clock_out_success": "下班打卡成功！時間",
        "auto_record_hint": "⚡ 點擊按鈕將自動記錄當前電腦時間",
        "today_records": "📋 當日個人打卡紀錄",
        "no_today_records": "今日尚無打卡資料。",
        "col_emp": "員工編號",
        "col_type": "打卡類型",
        "col_time": "時間",
        "type_in": "上班",
        "type_out": "下班",
        "tools_title": "🛠️ 修正 / 補登打卡紀錄 (Edit / Input Manual)",
        "tab_edit": "修改既有紀錄",
        "tab_manual": "手動補登紀錄",
        "select_record_to_edit": "選擇要修改的紀錄",
        "time_input_label": "請輸入時間 (格式: HH:MM，如 08:30)",
        "btn_save": "儲存修改",
        "btn_delete": "刪除",
        "msg_updated": "紀錄已更新！",
        "msg_deleted": "紀錄已刪除！",
        "btn_manual_add": "➕ 確認補登",
        "msg_manual_added": "補登成功！",
        "time_format_error": "❌ 時間格式錯誤！請輸入 24 小時制格式 (例如 08:35 或 17:30)",
        "warn_no_clock_in": "⚠️ 提醒：今日尚未記錄上班時間，是否確定要打下班卡？",
        "warn_under_hours": "⚠️ 提醒：今日於 {in_time} 上班，工時需滿 9 小時（合規下班時間為 {target_time}），目前工作時數未滿！是否確定打下班卡？",
        "btn_confirm_ok": "✔ OK (確認打卡)",
        "btn_confirm_cancel": "❌ 不OK (取消/不寫入)",
        "msg_action_cancelled": "已取消下班打卡，未寫入任何紀錄。"
    },
    "id": {
        "title": "Sistem Absensi Karyawan",
        "date_mgmt": "📅 Tanggal & Manajemen",
        "select_date": "Pilih Tanggal Absensi",
        "current_view_date": "Tanggal yang Dilihat",
        "export_title": "📤 Ekspor Data Absensi",
        "export_range": "Rentang Ekspor",
        "range_day": "Catatan Hari Ini",
        "range_month": "Catatan Bulan Ini",
        "range_all": "Semua Catatan",
        "download_excel": "📥 Unduh Excel (Format Impor)",
        "no_export_data": "Tidak ada data untuk diekspor.",
        "sys_version": "Versi Sistem",
        "lang_label": "🌐 Pilih Bahasa (語言)",
        "step1_select_emp": "1. Pilih ID Karyawan (請選擇員工編號)",
        "current_selected": "Karyawan Terpilih",
        "step2_clock": "2. Tekan Tombol Absensi (立即打卡)",
        "clock_in": "🟢 Masuk (Clock-In)",
        "clock_out": "🔴 Pulang (Clock-Out)",
        "clock_in_success": "Absensi Masuk Berhasil! Waktu",
        "clock_out_success": "Absensi Pulang Berhasil! Waktu",
        "auto_record_hint": "⚡ Tombol akan mencatat waktu komputer saat ini",
        "today_records": "📋 Catatan Absensi Pribadi Hari Ini",
        "no_today_records": "Belum ada catatan absensi hari ini.",
        "col_emp": "ID Karyawan",
        "col_type": "Tipe Absen",
        "col_time": "Waktu",
        "type_in": "Masuk",
        "type_out": "Pulang",
        "tools_title": "🛠️ Ubah / Tambah Data Manual (修正/補登)",
        "tab_edit": "Ubah Catatan",
        "tab_manual": "Input Manual",
        "select_record_to_edit": "Pilih Catatan untuk Diubah",
        "time_input_label": "Masukkan Waktu (Format: HH:MM, contoh: 08:30)",
        "btn_save": "Simpan Perubahan",
        "btn_delete": "Hapus",
        "msg_updated": "Catatan berhasil diperbarui!",
        "msg_deleted": "Catatan berhasil dihapus!",
        "btn_manual_add": "➕ Tambah Manual",
        "msg_manual_added": "Input manual berhasil!",
        "time_format_error": "❌ Format waktu salah! Harap masukkan format 24 jam (misal 08:35 atau 17:30)",
        "warn_no_clock_in": "⚠️ Peringatan: Belum ada catatan jam masuk hari ini. Apakah Anda yakin ingin absen pulang?",
        "warn_under_hours": "⚠️ Peringatan: Jam masuk hari ini pukul {in_time}, harus mencapai 9 jam (jam pulang standar {target_time}). Jam kerja belum cukup! Apakah yakin ingin pulang sekarang?",
        "btn_confirm_ok": "✔ OK (Tetap Absen Pulang)",
        "btn_confirm_cancel": "❌ Tidak / Batal (Batalkan)",
        "msg_action_cancelled": "Absen pulang dibatalkan, data tidak disimpan."
    }
}

# ==================== 頁面基本設定 ====================
st.set_page_config(
    page_title=f"外籍移工出勤打卡系統 {APP_VERSION}", 
    page_icon="⏰", 
    layout="wide"
)

# Session State 初始化
if "lang" not in st.session_state:
    st.session_state.lang = "zh"
if "selected_emp" not in st.session_state:
    st.session_state.selected_emp = EMPLOYEES[0]
if "pending_clock_out" not in st.session_state:
    st.session_state.pending_clock_out = None

def t(key):
    return TRANSLATIONS[st.session_state.lang].get(key, key)

# ==================== 工具函數：時間檢核與工時計算 ====================
def validate_time_format(time_str):
    if not time_str:
        return None
    time_str = time_str.strip()
    match = re.match(r"^([01]?\d|2[0-3]):([0-5]\d)$", time_str)
    if not match:
        return None
    h, m = int(match.group(1)), int(match.group(2))
    return f"{h:02d}:{m:02d}"

def calculate_target_checkout(in_time_str):
    h, m = map(int, in_time_str.split(":"))
    in_mins = h * 60 + m
    out_mins = in_mins + 9 * 60
    out_h = (out_mins // 60) % 24
    out_m = out_mins % 60
    return f"{out_h:02d}:{out_m:02d}"

def is_work_duration_sufficient(in_time_str, out_time_str):
    h1, m1 = map(int, in_time_str.split(":"))
    h2, m2 = map(int, out_time_str.split(":"))
    mins1 = h1 * 60 + m1
    mins2 = h2 * 60 + m2
    return (mins2 - mins1) >= (9 * 60)

# ==================== 資料庫初始化與操作 ====================
def get_db_connection():
    conn = sqlite3.connect(DB_FILE, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS punch_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            emp_id TEXT NOT NULL,
            date TEXT NOT NULL,       -- YYYY/MM/DD
            time TEXT NOT NULL,       -- HH:MM
            type TEXT NOT NULL,       -- 上班 / 下班
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.commit()
    conn.close()

init_db()

def add_record(emp_id, date_str, time_str, punch_type):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute(
        "INSERT INTO punch_records (emp_id, date, time, type) VALUES (?, ?, ?, ?)",
        (emp_id, date_str, time_str, punch_type)
    )
    conn.commit()
    conn.close()

def get_emp_records_by_date(emp_id, date_str):
    conn = get_db_connection()
    df = pd.read_sql_query(
        "SELECT id, emp_id, date, time, type FROM punch_records WHERE emp_id = ? AND date = ? ORDER BY time ASC",
        conn, params=(emp_id, date_str)
    )
    conn.close()
    return df

def get_all_records_by_date(date_str):
    conn = get_db_connection()
    df = pd.read_sql_query(
        "SELECT id, emp_id, date, time, type FROM punch_records WHERE date = ? ORDER BY emp_id ASC, time ASC",
        conn, params=(date_str,)
    )
    conn.close()
    return df

def get_all_records(start_date=None, end_date=None):
    conn = get_db_connection()
    if start_date and end_date:
        query = "SELECT id, emp_id, date, time, type FROM punch_records WHERE date BETWEEN ? AND ? ORDER BY date ASC, time ASC"
        df = pd.read_sql_query(query, conn, params=(start_date, end_date))
    else:
        query = "SELECT id, emp_id, date, time, type FROM punch_records ORDER BY date ASC, time ASC"
        df = pd.read_sql_query(query, conn)
    conn.close()
    return df

def update_record(record_id, new_time, new_type):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("UPDATE punch_records SET time = ?, type = ? WHERE id = ?", (new_time, new_type, record_id))
    conn.commit()
    conn.close()

def delete_record(record_id):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM punch_records WHERE id = ?", (record_id,))
    conn.commit()
    conn.close()

# ==================== Excel 匯出函數 ====================
def generate_excel_export(df_records):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "工作表1"

    ws.cell(row=1, column=1, value="打卡匯入")
    
    headers = ["序號", "員工編號", "日期", "時間"]
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=2, column=col_idx, value=h)
        
    rules = ["填寫規則", "必填", "必填", "非必填"]
    for col_idx, r in enumerate(rules, 1):
        ws.cell(row=3, column=col_idx, value=r)
        
    formats = ["欄位格式", "", "YYYY/MM/DD", "HH:MM"]
    for col_idx, f in enumerate(formats, 1):
        ws.cell(row=4, column=col_idx, value=f)

    for idx, row in df_records.iterrows():
        r_num = 5 + idx
        ws.cell(row=r_num, column=1, value=idx + 1)
        ws.cell(row=r_num, column=2, value=str(row["emp_id"]))
        ws.cell(row=r_num, column=3, value=str(row["date"]))
        ws.cell(row=r_num, column=4, value=str(row["time"]))

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 16

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# ==================== 側邊欄：日期選擇、語系與匯出 ====================
with st.sidebar:
    lang_choice = st.radio(
        t("lang_label"),
        options=["zh", "id"],
        format_func=lambda x: "🇹🇼 繁體中文" if x == "zh" else "🇮🇩 Bahasa Indonesia",
        index=0 if st.session_state.lang == "zh" else 1,
        horizontal=True
    )
    if lang_choice != st.session_state.lang:
        st.session_state.lang = lang_choice
        st.rerun()

    st.markdown("---")
    st.header(t("date_mgmt"))
    selected_date = st.date_input(t("select_date"), datetime.date.today())
    selected_date_str = selected_date.strftime("%Y/%m/%d")
    
    st.info(f"{t('current_view_date')}：**{selected_date_str}**")
    st.markdown("---")
    
    st.subheader(t("export_title"))
    range_map = {
        t("range_day"): "day",
        t("range_month"): "month",
        t("range_all"): "all"
    }
    export_range_label = st.selectbox(t("export_range"), list(range_map.keys()))
    export_range = range_map[export_range_label]
    
    if export_range == "day":
        export_df = get_all_records_by_date(selected_date_str)
    elif export_range == "month":
        first_day = selected_date.replace(day=1).strftime("%Y/%m/%d")
        next_month = selected_date.replace(day=28) + datetime.timedelta(days=4)
        last_day = (next_month - datetime.timedelta(days=next_month.day)).strftime("%Y/%m/%d")
        export_df = get_all_records(first_day, last_day)
    else:
        export_df = get_all_records()
        
    if not export_df.empty:
        excel_data = generate_excel_export(export_df)
        st.download_button(
            label=t("download_excel"),
            data=excel_data,
            file_name=f"打卡匯出_{selected_date_str.replace('/', '')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    else:
        st.caption(t("no_export_data"))

    st.markdown("---")
    st.caption(f"{t('sys_version')}: `{APP_VERSION}`")

# ==================== 主要介面 ====================
st.title(f"⏰ {t('title')} ({APP_VERSION})")
st.markdown("---")

# 1. 選擇員工編號
st.subheader(t("step1_select_emp"))
emp_cols = st.columns(len(EMPLOYEES))
for i, emp in enumerate(EMPLOYEES):
    btn_type = "primary" if st.session_state.selected_emp == emp else "secondary"
    if emp_cols[i].button(f"👤 {emp}", key=f"emp_btn_{emp}", type=btn_type, use_container_width=True):
        st.session_state.selected_emp = emp
        st.session_state.pending_clock_out = None  # 切換人員時重設待確認狀態
        st.rerun()

current_emp = st.session_state.selected_emp
st.success(f"{t('current_selected')}：**{current_emp}**")

# 2. 工時不足確認卡片 (若觸發二次防呆)
if st.session_state.pending_clock_out and st.session_state.pending_clock_out["emp_id"] == current_emp:
    pending_info = st.session_state.pending_clock_out
    with st.container():
        st.error(pending_info["warn_msg"])
        col_ok, col_cancel = st.columns(2)
        with col_ok:
            if st.button(t("btn_confirm_ok"), type="primary", use_container_width=True):
                # 確認打卡 -> 正式寫入 DB
                add_record(pending_info["emp_id"], pending_info["date"], pending_info["time"], "下班")
                st.toast(f"✅ {pending_info['emp_id']} {t('clock_out_success')}：{pending_info['time']}", icon="👋")
                st.session_state.pending_clock_out = None
                st.rerun()
        with col_cancel:
            if st.button(t("btn_confirm_cancel"), type="secondary", use_container_width=True):
                # 取消打卡 -> 不寫入 DB
                st.info(t("msg_action_cancelled"))
                st.session_state.pending_clock_out = None
                st.rerun()
    st.markdown("---")

# 3. 即時打卡按鈕
st.subheader(t("step2_clock"))
clock_cols = st.columns(2)

now = datetime.datetime.now()
current_time_str = now.strftime("%H:%M")

emp_today_records = get_emp_records_by_date(current_emp, selected_date_str)

with clock_cols[0]:
    if st.button(t("clock_in"), use_container_width=True, type="primary"):
        st.session_state.pending_clock_out = None
        add_record(current_emp, selected_date_str, current_time_str, "上班")
        st.toast(f"✅ {current_emp} {t('clock_in_success')}：{current_time_str}", icon="🎉")
        st.rerun()

with clock_cols[1]:
    if st.button(t("clock_out"), use_container_width=True):
        clock_in_records = emp_today_records[emp_today_records["type"] == "上班"]
        
        # 情況 A: 當天完全沒有打上班卡 -> 觸發確認防呆
        if clock_in_records.empty:
            st.session_state.pending_clock_out = {
                "emp_id": current_emp,
                "date": selected_date_str,
                "time": current_time_str,
                "warn_msg": t("warn_no_clock_in")
            }
            st.rerun()
        else:
            first_in_time = clock_in_records.iloc[0]["time"]
            target_out_time = calculate_target_checkout(first_in_time)
            
            # 情況 B: 未滿 9 小時 -> 暫緩寫入，跳出確認選項
            if not is_work_duration_sufficient(first_in_time, current_time_str):
                st.session_state.pending_clock_out = {
                    "emp_id": current_emp,
                    "date": selected_date_str,
                    "time": current_time_str,
                    "warn_msg": t("warn_under_hours").format(
                        in_time=first_in_time, 
                        target_time=target_out_time
                    )
                }
                st.rerun()
            else:
                # 情況 C: 工時已滿 9 小時 -> 直接寫入
                st.session_state.pending_clock_out = None
                add_record(current_emp, selected_date_str, current_time_str, "下班")
                st.toast(f"✅ {current_emp} {t('clock_out_success')}：{current_time_str}", icon="👋")
                st.rerun()

st.caption(f"{t('auto_record_hint')}：**{current_time_str}**")

st.markdown("---")

# 4. 當日個人打卡紀錄 (隱私隔離)
st.subheader(f"{t('today_records')} - {current_emp} ({selected_date_str})")

if not emp_today_records.empty:
    display_df = emp_today_records[["emp_id", "type", "time"]].copy()
    display_df["type"] = display_df["type"].apply(lambda x: t("type_in") if x == "上班" else t("type_out"))
    display_df = display_df.rename(columns={
        "emp_id": t("col_emp"),
        "type": t("col_type"),
        "time": t("col_time")
    })
    st.dataframe(display_df, use_container_width=True, hide_index=True)
else:
    st.info(f"{current_emp} {t('no_today_records')}")

# 5. 補登與編輯管理區
with st.expander(t("tools_title")):
    tab_edit, tab_manual = st.tabs([t("tab_edit"), t("tab_manual")])
    
    with tab_edit:
        if not emp_today_records.empty:
            record_options = {
                f"ID {r['id']} - ({r['type']} {r['time']})": r['id']
                for _, r in emp_today_records.iterrows()
            }
            selected_record_label = st.selectbox(t("select_record_to_edit"), list(record_options.keys()))
            selected_record_id = record_options[selected_record_label]
            target_row = emp_today_records[emp_today_records["id"] == selected_record_id].iloc[0]

            col_edit1, col_edit2, col_edit3 = st.columns(3)
            with col_edit1:
                edit_type = st.selectbox(
                    t("col_type"), 
                    ["上班", "下班"], 
                    format_func=lambda x: t("type_in") if x == "上班" else t("type_out"),
                    index=0 if target_row["type"] == "上班" else 1
                )
            with col_edit2:
                edit_time_input = st.text_input(t("time_input_label"), value=str(target_row["time"]), key="edit_time_key")
            with col_edit3:
                st.write("")
                st.write("")
                col_btn1, col_btn2 = st.columns(2)
                if col_btn1.button(t("btn_save"), use_container_width=True, type="primary"):
                    valid_time = validate_time_format(edit_time_input)
                    if valid_time:
                        update_record(selected_record_id, valid_time, edit_type)
                        st.success(t("msg_updated"))
                        st.rerun()
                    else:
                        st.error(t("time_format_error"))
                if col_btn2.button(t("btn_delete"), type="secondary", use_container_width=True):
                    delete_record(selected_record_id)
                    st.warning(t("msg_deleted"))
                    st.rerun()
        else:
            st.caption(t("no_today_records"))
            
    with tab_manual:
        col_m1, col_m2, col_m3 = st.columns(3)
        with col_m1:
            manual_emp = st.selectbox(t("col_emp"), EMPLOYEES, index=EMPLOYEES.index(current_emp), key="manual_emp")
        with col_m2:
            manual_type = st.selectbox(
                t("col_type"), 
                ["上班", "下班"], 
                format_func=lambda x: t("type_in") if x == "上班" else t("type_out"),
                key="manual_type"
            )
        with col_m3:
            manual_time_input = st.text_input(t("time_input_label"), value=current_time_str, key="manual_time_key")
            
        if st.button(t("btn_manual_add"), type="primary"):
            valid_time = validate_time_format(manual_time_input)
            if valid_time:
                add_record(manual_emp, selected_date_str, valid_time, manual_type)
                st.success(f"{t('msg_manual_added')} ({manual_emp} - {manual_type} {valid_time})")
                st.rerun()
            else:
                st.error(t("time_format_error"))
